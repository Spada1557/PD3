from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_
import os
import uuid
import csv
import io
from openpyxl import load_workbook
from app.database import get_db
from app.models import Product, Category, Unit, Stock, StockMovement
from app.schemas import ProductCreate, ProductUpdate, ProductResponse, PaginatedResponse
from app.api.deps import get_current_user
from app.config import settings

router = APIRouter(prefix="/api/v1/products", tags=["products"])


@router.get("", response_model=PaginatedResponse)
def list_products(
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    search: str = None,
    category_id: int = None,
    status: str = None,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    q = db.query(Product).options(joinedload(Product.category), joinedload(Product.unit))
    if search:
        q = q.filter(or_(
            Product.name.ilike(f"%{search}%"),
            Product.article.ilike(f"%{search}%")
        ))
    if category_id:
        q = q.filter(Product.category_id == category_id)
    if status:
        q = q.filter(Product.status == status)
    total = q.count()
    items = q.offset((page - 1) * per_page).limit(per_page).all()
    return PaginatedResponse(total=total, page=page, per_page=per_page, data=[ProductResponse.model_validate(p) for p in items])


@router.get("/{product_id}", response_model=ProductResponse)
def get_product(product_id: int, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    p = db.query(Product).options(joinedload(Product.category), joinedload(Product.unit)).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail={"error": "not_found", "message": "Товар не найден"})
    return p


@router.post("", response_model=ProductResponse, status_code=201)
def create_product(body: ProductCreate, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    existing = db.query(Product).filter(Product.article == body.article).first()
    if existing:
        raise HTTPException(status_code=400, detail={"error": "duplicate", "message": "Товар с таким артикулом уже существует"})
    cat = db.query(Category).filter(Category.id == body.category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail={"error": "not_found", "message": "Категория не найдена"})
    unit = db.query(Unit).filter(Unit.id == body.unit_id).first()
    if not unit:
        raise HTTPException(status_code=404, detail={"error": "not_found", "message": "Единица измерения не найдена"})
    p = Product(**body.model_dump())
    db.add(p)
    db.commit()
    db.refresh(p)
    p = db.query(Product).options(joinedload(Product.category), joinedload(Product.unit)).filter(Product.id == p.id).first()
    return p


@router.put("/{product_id}", response_model=ProductResponse)
def update_product(product_id: int, body: ProductUpdate, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail={"error": "not_found", "message": "Товар не найден"})
    if body.article and body.article != p.article:
        existing = db.query(Product).filter(Product.article == body.article, Product.id != product_id).first()
        if existing:
            raise HTTPException(status_code=400, detail={"error": "duplicate", "message": "Товар с таким артикулом уже существует"})
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(p, field, value)
    db.commit()
    db.refresh(p)
    p = db.query(Product).options(joinedload(Product.category), joinedload(Product.unit)).filter(Product.id == p.id).first()
    return p


@router.delete("/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail={"error": "not_found", "message": "Товар не найден"})
    has_movements = db.query(StockMovement).filter(StockMovement.product_id == product_id).count() > 0
    if has_movements:
        raise HTTPException(status_code=400, detail={"error": "has_movements", "message": "Нельзя удалить товар — по нему были движения на складе"})
    stocks = db.query(Stock).filter(Stock.product_id == product_id).all()
    for s in stocks:
        db.delete(s)
    db.delete(p)
    db.commit()
    return {"message": "Товар удалён"}


@router.post("/{product_id}/upload-image")
def upload_image(product_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail={"error": "not_found", "message": "Товар не найден"})
    if file.content_type not in ("image/jpeg", "image/png"):
        raise HTTPException(status_code=400, detail={"error": "invalid_format", "message": "Допустимы только JPG и PNG"})
    content = file.file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail={"error": "too_large", "message": "Файл превышает 2 МБ"})
    ext = "jpg" if file.content_type == "image/jpeg" else "png"
    filename = f"{uuid.uuid4()}.{ext}"
    filepath = os.path.join(settings.UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        f.write(content)
    p.image_path = filename
    db.commit()
    return {"image_path": filename}


@router.post("/import")
def import_products(file: UploadFile = File(...), db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    content = file.file.read()
    filename = file.filename or ""

    if filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")), delimiter=";")
    elif filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2):
            row_dict = {headers[i]: (cell.value or "") for i, cell in enumerate(row)}
            rows.append(row_dict)
        reader = rows
    else:
        raise HTTPException(status_code=400, detail={"error": "invalid_format", "message": "Поддерживаются только .xlsx и .csv"})

    imported = 0
    errors = []
    for i, row in enumerate(reader):
        try:
            name = str(row.get("name", row.get("Наименование", ""))).strip()
            article = str(row.get("article", row.get("Артикул", ""))).strip()
            price = float(row.get("price", row.get("Цена", 0)))
            cost = float(row.get("cost", row.get("Себестоимость", 0)))
            min_stock = float(row.get("min_stock", row.get("МинОстаток", 10)))
            cat_name = str(row.get("category", row.get("Категория", ""))).strip()
            unit_name = str(row.get("unit", row.get("ЕдИзм", "шт"))).strip()

            if not name or not article:
                errors.append(f"Строка {i+2}: отсутствует наименование или артикул")
                continue

            existing = db.query(Product).filter(Product.article == article).first()
            if existing:
                errors.append(f"Строка {i+2}: артикул {article} уже существует")
                continue

            category = db.query(Category).filter(Category.name == cat_name).first()
            if not category and cat_name:
                category = Category(name=cat_name)
                db.add(category)
                db.flush()

            unit = db.query(Unit).filter(Unit.name == unit_name).first() or db.query(Unit).first()
            if not unit and unit_name:
                unit = Unit(name=unit_name, short_name=unit_name[:10])
                db.add(unit)
                db.flush()

            p = Product(
                name=name, article=article,
                category_id=category.id if category else 1,
                unit_id=unit.id if unit else 1,
                price=price, cost=cost, min_stock=min_stock,
            )
            db.add(p)
            imported += 1
        except Exception as e:
            errors.append(f"Строка {i+2}: {str(e)}")

    db.commit()
    return {"imported": imported, "errors": errors}
