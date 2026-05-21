from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func as sa_func
from app.database import get_db
from app.models import Sale, SaleItem, Product, Category, Stock, Purchase
from app.schemas import ReportRequest, PaginatedResponse
from app.api.deps import get_current_user, require_role
import io
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/v1/reports", tags=["reports"])


@router.get("/sales", response_model=PaginatedResponse)
def sales_report(
    date_from: str = None,
    date_to: str = None,
    category_id: int = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    user: dict = Depends(require_role("admin", "manager")),
):
    q = db.query(Sale).options(
        joinedload(Sale.client),
        joinedload(Sale.items).joinedload(SaleItem.product).joinedload(Product.category)
    ).filter(Sale.status == "posted")

    if date_from:
        q = q.filter(Sale.date >= date_from)
    if date_to:
        q = q.filter(Sale.date <= date_to)
    if category_id:
        q = q.join(Sale.items).join(SaleItem.product).filter(Product.category_id == category_id)

    total = q.count()
    items = q.offset((page - 1) * per_page).limit(per_page).all()

    data = []
    for s in items:
        d = {
            "id": s.id,
            "number": s.number,
            "date": s.date,
            "client": s.client.name if s.client else "",
            "total_amount": s.total_amount,
            "total_cost": s.total_cost,
            "total_profit": s.total_profit,
            "margin": round(s.total_profit / s.total_amount * 100, 1) if s.total_amount > 0 else 0,
            "items_count": len(s.items),
        }
        data.append(d)

    return PaginatedResponse(total=total, page=page, per_page=per_page, data=data)


@router.get("/stock-snapshot")
def stock_snapshot(
    snapshot_date: str = None,
    warehouse_id: int = None,
    category_id: int = None,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    q = db.query(Stock).options(
        joinedload(Stock.product).joinedload(Product.category),
        joinedload(Stock.warehouse),
    )
    if warehouse_id:
        q = q.filter(Stock.warehouse_id == warehouse_id)
    if category_id:
        q = q.join(Product).filter(Product.category_id == category_id)

    data = []
    total_value = 0.0
    for s in q.all():
        available = s.quantity - s.reserved
        value = available * s.avg_cost
        total_value += value
        data.append({
            "product_id": s.product_id,
            "product_name": s.product.name if s.product else "",
            "article": s.product.article if s.product else "",
            "category": s.product.category.name if s.product and s.product.category else "",
            "warehouse": s.warehouse.name if s.warehouse else "",
            "quantity": s.quantity,
            "reserved": s.reserved,
            "available": available,
            "avg_cost": s.avg_cost,
            "value": round(value, 2),
        })

    return {"data": data, "total_value": round(total_value, 2), "snapshot_date": snapshot_date or str(date.today())}


@router.get("/top-products")
def top_products(
    date_from: str = None,
    date_to: str = None,
    limit: int = Query(10, ge=1, le=100),
    order_by: str = Query("revenue", pattern="^(revenue|quantity)$"),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    q = db.query(
        SaleItem.product_id,
        Product.name,
        Product.article,
        sa_func.sum(SaleItem.quantity).label("total_quantity"),
        sa_func.sum(SaleItem.amount).label("total_revenue"),
        sa_func.sum(SaleItem.cost).label("total_cost"),
        sa_func.sum(SaleItem.amount - SaleItem.cost).label("total_profit"),
    ).join(Sale).join(Product).filter(Sale.status == "posted")

    if date_from:
        q = q.filter(Sale.date >= date_from)
    if date_to:
        q = q.filter(Sale.date <= date_to)

    q = q.group_by(SaleItem.product_id)

    if order_by == "revenue":
        q = q.order_by(sa_func.sum(SaleItem.amount).desc())
    else:
        q = q.order_by(sa_func.sum(SaleItem.quantity).desc())

    results = q.limit(limit).all()
    data = [{
        "product_id": r.product_id,
        "product_name": r.name,
        "article": r.article,
        "total_quantity": round(r.total_quantity, 2),
        "total_revenue": round(r.total_revenue, 2),
        "total_cost": round(r.total_cost, 2),
        "total_profit": round(r.total_profit or 0, 2),
        "margin": round((r.total_profit or 0) / r.total_revenue * 100, 1) if r.total_revenue > 0 else 0,
    } for r in results]

    return {"data": data}


@router.get("/purchases", response_model=PaginatedResponse)
def purchases_report(
    date_from: str = None,
    date_to: str = None,
    supplier_id: int = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
    user: dict = Depends(require_role("admin", "warehouse")),
):
    q = db.query(Purchase).options(
        joinedload(Purchase.supplier),
        joinedload(Purchase.items),
    ).filter(Purchase.status == "posted")

    if date_from:
        q = q.filter(Purchase.date >= date_from)
    if date_to:
        q = q.filter(Purchase.date <= date_to)
    if supplier_id:
        q = q.filter(Purchase.supplier_id == supplier_id)

    total = q.count()
    items = q.offset((page - 1) * per_page).limit(per_page).all()

    data = [{
        "id": p.id,
        "number": p.number,
        "date": p.date,
        "supplier": p.supplier.name if p.supplier else "",
        "total_amount": p.total_amount,
        "items_count": len(p.items),
    } for p in items]

    return PaginatedResponse(total=total, page=page, per_page=per_page, data=data)


@router.get("/export/sales")
def export_sales_excel(
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_role("admin", "manager")),
):
    q = db.query(Sale).options(
        joinedload(Sale.client),
        joinedload(Sale.items).joinedload(SaleItem.product).joinedload(Product.category)
    ).filter(Sale.status == "posted")

    if date_from:
        q = q.filter(Sale.date >= date_from)
    if date_to:
        q = q.filter(Sale.date <= date_to)

    sales = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Продажи"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["№ документа", "Дата", "Клиент", "Товар", "Артикул", "Кол-во", "Цена", "Сумма", "Себестоимость", "Прибыль"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for s in sales:
        for item in s.items:
            ws.cell(row=row, column=1, value=s.number).border = thin_border
            ws.cell(row=row, column=2, value=s.date.strftime("%d.%m.%Y") if s.date else "").border = thin_border
            ws.cell(row=row, column=3, value=s.client.name if s.client else "").border = thin_border
            ws.cell(row=row, column=4, value=item.product.name if item.product else "").border = thin_border
            ws.cell(row=row, column=5, value=item.product.article if item.product else "").border = thin_border
            ws.cell(row=row, column=6, value=item.quantity).border = thin_border
            ws.cell(row=row, column=7, value=item.price).border = thin_border
            ws.cell(row=row, column=8, value=item.amount).border = thin_border
            ws.cell(row=row, column=9, value=item.cost).border = thin_border
            ws.cell(row=row, column=10, value=round(item.amount - item.cost, 2)).border = thin_border
            row += 1

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/stock")
def export_stock_excel(
    warehouse_id: int = None,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    q = db.query(Stock).options(
        joinedload(Stock.product).joinedload(Product.category),
        joinedload(Stock.warehouse),
    )
    if warehouse_id:
        q = q.filter(Stock.warehouse_id == warehouse_id)

    stocks = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Товар", "Артикул", "Категория", "Склад", "Количество", "Зарезервировано", "Доступно", "Ср. себестоимость", "Стоимость остатка"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row, s in enumerate(stocks, 2):
        ws.cell(row=row, column=1, value=s.product.name if s.product else "").border = thin_border
        ws.cell(row=row, column=2, value=s.product.article if s.product else "").border = thin_border
        ws.cell(row=row, column=3, value=s.product.category.name if s.product and s.product.category else "").border = thin_border
        ws.cell(row=row, column=4, value=s.warehouse.name if s.warehouse else "").border = thin_border
        ws.cell(row=row, column=5, value=s.quantity).border = thin_border
        ws.cell(row=row, column=6, value=s.reserved).border = thin_border
        ws.cell(row=row, column=7, value=s.quantity - s.reserved).border = thin_border
        ws.cell(row=row, column=8, value=s.avg_cost).border = thin_border
        ws.cell(row=row, column=9, value=round((s.quantity - s.reserved) * s.avg_cost, 2)).border = thin_border

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
